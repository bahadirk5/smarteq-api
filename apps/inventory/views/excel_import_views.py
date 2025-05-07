from rest_framework import status
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from openpyxl import Workbook
from apps.common.responses import success_response, error_response
from apps.inventory.services.excel_import_service import ExcelImportService
from apps.inventory.repositories.excel_import_repository import ExcelImportRepository
from apps.inventory.serializers.excel_import_serializer import (
    ExcelImportSerializer,
    ExcelImportCreateSerializer,
    ExcelTemplateInfoSerializer
)
import logging

logger = logging.getLogger(__name__)

class ExcelImportListCreateView(APIView):
    """
    API endpoint for listing and creating Excel imports
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """
        Get all Excel imports
        """
        imports, error = ExcelImportService.get_imports()
        if error:
            return error_response(error, status.HTTP_400_BAD_REQUEST)
        
        serializer = ExcelImportSerializer(imports, many=True)
        return success_response(serializer.data)
    
    def post(self, request):
        """
        Create a new Excel import
        """
        serializer = ExcelImportCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return error_response(serializer.errors, status.HTTP_400_BAD_REQUEST)
        
        # Proje ID'yi al - önce serializer içinden (varsa), sonra request.data'dan (camelCase uyumluluk için)
        project_id = serializer.validated_data.get('project_id') or request.data.get('projectId') or request.data.get('project_id')
        
        excel_import, error = ExcelImportService.create_import(
            import_type=serializer.validated_data['import_type'],
            file=serializer.validated_data['file'],
            notes=serializer.validated_data.get('notes'),
            processed_by=request.user,
            project_id=project_id
        )
        
        if error:
            return error_response(error, status.HTTP_400_BAD_REQUEST)
        
        result_serializer = ExcelImportSerializer(excel_import)
        return success_response(result_serializer.data, status.HTTP_201_CREATED)


class ExcelImportDetailView(APIView):
    """
    API endpoint for retrieving and processing an Excel import
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, import_id):
        """
        Get an Excel import by ID
        """
        excel_import, error = ExcelImportService.get_import_by_id(import_id)
        if error:
            return error_response(error, status.HTTP_404_NOT_FOUND)
        
        serializer = ExcelImportSerializer(excel_import)
        return success_response(serializer.data)
    
    def post(self, request, import_id):
        """
        Process an Excel import
        """
        # Bu noktada project_id gelmiş mi diye kontrol ediyoruz (hem camelCase hem snake_case destekle)
        project_id = request.data.get('projectId') or request.data.get('project_id')
        
        # Eğer geldiyse Excel import'un notlarını güncelliyoruz
        if project_id:
            # Mimariye uygun olarak Service katmanını kullanıyoruz, doğrudan Repository değil
            updated, error = ExcelImportService.update_import_notes_with_project_id(import_id, project_id)
            if error:
                return error_response(error, status.HTTP_400_BAD_REQUEST)
        
        excel_import, error = ExcelImportService.get_import_by_id(import_id)
        if error:
            return error_response(error, status.HTTP_404_NOT_FOUND)
        
        if excel_import.import_type == 'RAW_MATERIALS':
            result, error = ExcelImportService.process_raw_materials_import(import_id)
        elif excel_import.import_type == 'BOM':
            result, error = ExcelImportService.process_bom_import(import_id)
        elif excel_import.import_type == 'ELECTRONIC_COMPONENTS':
            result, error = ExcelImportService.process_electronic_components_import(import_id)
        elif excel_import.import_type == 'PRODUCTS':
            result, error = ExcelImportService.process_products_import(import_id)
        else:
            return error_response('Unsupported import type', status.HTTP_400_BAD_REQUEST)
        
        if error:
            return error_response(error, status.HTTP_400_BAD_REQUEST)
        
        # Get updated import data
        updated_import, _ = ExcelImportService.get_import_by_id(import_id)
        serializer = ExcelImportSerializer(updated_import)
        
        return success_response(serializer.data)


class ExcelTemplateView(APIView):
    """
    API endpoint for downloading Excel templates
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, import_type):
        """
        Download an Excel template for the specified import type
        """
        if import_type not in ['RAW_MATERIALS', 'BOM']:
            return error_response('Invalid import type', status.HTTP_400_BAD_REQUEST)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        
        if import_type == 'RAW_MATERIALS':
            # Set headers for raw materials template
            headers = ['sku', 'name', 'description', 'category', 'unit_of_measure', 
                      'quantity', 'purchase_price', 'purchase_date', 'supplier', 'invoice_reference']
            ws.title = 'Raw Materials'
            
            # Add a sample row
            sample_data = [
                'RM-001', 'Sample Raw Material', 'Description here', 'Chemicals', 'kg',
                '100', '15.50', '2023-01-15', 'Supplier Name', 'INV-12345'
            ]
        else:  # BOM template
            # Set headers for BOM template
            headers = ['output_sku', 'input_sku', 'quantity_required', 'unit_of_measure', 
                      'sequence', 'is_optional', 'is_default', 'alternative_group']
            ws.title = 'Bill of Materials'
            
            # Add sample rows demonstrating alternatives
            sample_data = [
                ['FP-001', 'RM-001', '2', 'kg', '10', '', '', ''],
                ['FP-001', 'RM-002', '1', 'l', '20', '', '', ''],
                ['FP-001', 'RM-003', '0.5', 'kg', '30', 'Y', 'Y', 'GROUP1'],
                ['FP-001', 'RM-004', '0.5', 'kg', '30', 'Y', 'N', 'GROUP1']
            ]
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            # Add bold formatting
            cell.font = cell.font.copy(bold=True)
        
        # Write sample data
        if import_type == 'RAW_MATERIALS':
            for col_idx, value in enumerate(sample_data, 1):
                ws.cell(row=2, column=col_idx, value=value)
        else:
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={import_type.lower()}_template.xlsx'
        
        # Save workbook to response
        wb.save(response)
        
        return response


class ExcelTemplateInfoView(APIView):
    """
    API endpoint for getting information about Excel templates
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """
        Get information about available Excel templates
        """
        templates = [
            {
                'import_type': 'RAW_MATERIALS',
                'template_name': 'Raw Materials Import',
                'required_columns': ['sku', 'name', 'unit_of_measure'],
                'optional_columns': ['description', 'category', 'quantity', 'purchase_price',
                                    'purchase_date', 'supplier', 'invoice_reference'],
                'column_descriptions': {
                    'sku': 'Unique identifier for the item',
                    'name': 'Item name',
                    'description': 'Item description',
                    'category': 'Category name',
                    'unit_of_measure': 'Unit of measure (e.g., kg, l, piece)',
                    'quantity': 'Initial quantity',
                    'purchase_price': 'Purchase price per unit',
                    'purchase_date': 'Date of purchase (YYYY-MM-DD)',
                    'supplier': 'Supplier name',
                    'invoice_reference': 'Invoice reference'
                }
            },
            {
                'import_type': 'BOM',
                'template_name': 'Bill of Materials Import',
                'required_columns': ['output_sku', 'input_sku', 'quantity_required', 'unit_of_measure'],
                'optional_columns': ['sequence', 'is_optional', 'is_default', 'alternative_group'],
                'column_descriptions': {
                    'output_sku': 'SKU of the output item (final or intermediate product)',
                    'input_sku': 'SKU of the input item (component)',
                    'quantity_required': 'Quantity of input item required for one unit of output item',
                    'unit_of_measure': 'Unit of measure for the input item',
                    'sequence': 'Assembly sequence number (default: 10)',
                    'is_optional': 'Whether the component is optional (Y/N, default: N)',
                    'is_default': 'Whether this is the default component in an alternative group (Y/N, default: Y)',
                    'alternative_group': 'Group identifier for alternative components'
                }
            }
        ]
        
        serializer = ExcelTemplateInfoSerializer(templates, many=True)
        return success_response(serializer.data)
